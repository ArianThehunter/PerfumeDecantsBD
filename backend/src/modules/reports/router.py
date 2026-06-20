"""
Reports Module — API Router
==============================
Generate and download sales, inventory, and customer reports in PDF, Excel, or CSV format.
"""

import os
import csv
from datetime import UTC, datetime
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from sqlalchemy import select, func, desc

# Excel & PDF imports
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from src.core.dependencies import DbSession, DbSession as DbSessionDep, require_role, CurrentUser
from src.core.pagination import PaginatedResponse
from src.core.dependencies import Pagination
from src.modules.reports.models import Report
from src.modules.orders.models import Order
from src.modules.products.models import ProductVariant, Product
from src.modules.auth.models import User

router = APIRouter(
    prefix="/reports",
    tags=["Reports"],
    dependencies=[Depends(require_role("SuperAdmin", "Manager"))],
)

# ── Schemas ──────────────────────────────────────────────────────────────────

class ReportGenerateRequest(BaseModel):
    name: str = Field(..., max_length=255)
    report_type: str = Field(..., pattern="^(sales|revenue|profit|inventory|customer)$")
    format: str = Field(..., pattern="^(pdf|excel|csv)$")

class ReportResponse(BaseModel):
    id: int
    name: str
    report_type: str
    format: str
    status: str
    created_at: datetime

    model_config = {"from_attributes": True}

# ── Utility Generation Functions ──────────────────────────────────────────────

REPORTS_DIR = Path("./uploads/reports")
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

async def fetch_report_data(report_type: str, db: DbSession):
    """Query data from DB depending on the report type."""
    if report_type == "sales":
        # Get order records
        stmt = (
            select(
                Order.order_number,
                Order.status,
                Order.total,
                Order.created_at
            )
            .order_by(Order.created_at.desc())
            .limit(200)
        )
        res = await db.execute(stmt)
        return [
            ["Order Number", "Status", "Total Amount ($)", "Date Created"],
            *[[row.order_number, row.status, float(row.total), row.created_at.strftime("%Y-%m-%d %H:%M")] for row in res.all()]
        ]
        
    elif report_type == "inventory":
        # Get stock levels
        stmt = (
            select(
                Product.name.label("product_name"),
                ProductVariant.name.label("variant_name"),
                ProductVariant.sku,
                ProductVariant.stock_quantity,
                ProductVariant.price
            )
            .join(Product, ProductVariant.product_id == Product.id)
            .where(ProductVariant.is_active == True)  # noqa: E712
            .order_by(ProductVariant.stock_quantity)
            .limit(200)
        )
        res = await db.execute(stmt)
        return [
            ["Product Name", "Size/Concentration", "SKU", "Stock Qty", "Price ($)"],
            *[[row.product_name, row.variant_name, row.sku, row.stock_quantity, float(row.price)] for row in res.all()]
        ]
        
    elif report_type == "customer":
        # Get customer spending
        stmt = (
            select(
                User.first_name,
                User.last_name,
                User.email,
                func.count(Order.id).label("order_count"),
                func.coalesce(func.sum(Order.total), 0).label("total_spent")
            )
            .join(Order, Order.user_id == User.id, isouter=True)
            .group_by(User.id)
            .order_by(desc("total_spent"))
            .limit(200)
        )
        res = await db.execute(stmt)
        return [
            ["Customer Name", "Email Address", "Total Orders", "Total Spent ($)"],
            *[[f"{row.first_name} {row.last_name}", row.email, row.order_count, float(row.total_spent)] for row in res.all()]
        ]
        
    else:  # revenue/profit
        # Simple aggregated revenue metrics
        stmt = (
            select(
                func.date(Order.created_at).label("day"),
                func.count(Order.id).label("orders"),
                func.sum(Order.total).label("revenue")
            )
            .where(Order.status.in_(["confirmed", "processing", "shipped", "delivered"]))
            .group_by(func.date(Order.created_at))
            .order_by(desc("day"))
            .limit(100)
        )
        res = await db.execute(stmt)
        return [
            ["Date", "Orders Count", "Gross Revenue ($)"],
            *[[str(row.day), row.orders, float(row.revenue or 0)] for row in res.all()]
        ]

def generate_csv(filepath: Path, title: str, data: list):
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow([title])
        writer.writerow([])
        writer.writerows(data)

def generate_excel(filepath: Path, title: str, data: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Title style
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = openpyxl.styles.Font(size=14, bold=True)
    
    # Write rows
    start_row = 3
    for r_idx, row in enumerate(data):
        for c_idx, val in enumerate(row):
            cell = ws.cell(row=start_row + r_idx, column=c_idx + 1, value=val)
            # Headers style
            if r_idx == 0:
                cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                cell.fill = openpyxl.styles.PatternFill(start_color="C9A84C", end_color="C9A84C", fill_type="solid")
                
    wb.save(filepath)

def generate_pdf(filepath: Path, title: str, data: list):
    doc = SimpleDocTemplate(str(filepath), pagesize=letter)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="TitleStyle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        textColor=colors.HexColor("#C9A84C"),
        spaceAfter=15
    )
    
    # Add title
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 10))
    
    # Table layout
    # Keep row data simple and wrap with Paragraphs
    table_data = []
    for r_idx, row in enumerate(data):
        row_data = []
        for val in row:
            if r_idx == 0:
                # Header
                style = ParagraphStyle(name=f"H_{r_idx}", parent=styles["Normal"], fontName="Helvetica-Bold", textColor=colors.white)
            else:
                style = styles["Normal"]
            row_data.append(Paragraph(str(val), style))
        table_data.append(row_data)
        
    t = Table(table_data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#6B1D3A")),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    story.append(t)
    doc.build(story)

# ── Router Endpoints ─────────────────────────────────────────────────────────

@router.get("", response_model=PaginatedResponse[ReportResponse])
async def list_reports(pagination: Pagination, db: DbSessionDep):
    """List all generated business reports (admin)."""
    count = (await db.execute(select(func.count(Report.id)))).scalar() or 0
    stmt = (
        select(Report)
        .order_by(Report.created_at.desc())
        .offset(pagination.offset)
        .limit(pagination.page_size)
    )
    result = await db.execute(stmt)
    return PaginatedResponse.create(
        list(result.scalars().all()), count, pagination.page, pagination.page_size
    )

@router.post("/generate", response_model=ReportResponse, status_code=201)
async def generate_report(
    req: ReportGenerateRequest,
    user: CurrentUser,
    db: DbSessionDep
):
    """Request a new report generation of a specific type and format."""
    # Generate unique filename
    timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    extension = "xlsx" if req.format == "excel" else req.format
    filename = f"{req.report_type}_{timestamp}.{extension}"
    filepath = REPORTS_DIR / filename
    
    # Fetch data
    data = await fetch_report_data(req.report_type, db)
    
    # Run format-specific builder
    try:
        if req.format == "csv":
            generate_csv(filepath, req.name, data)
        elif req.format == "excel":
            generate_excel(filepath, req.name, data)
        elif req.format == "pdf":
            generate_pdf(filepath, req.name, data)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to generate document file: {str(e)}"
        )
        
    # Write metadata record to database
    report = Report(
        name=req.name,
        report_type=req.report_type,
        format=req.format,
        file_path=str(filepath),
        status="completed",
        created_by=f"{user.first_name} {user.last_name}"
    )
    
    db.add(report)
    await db.flush()
    return report

@router.get("/{report_id}/download")
async def download_report(report_id: int, db: DbSessionDep):
    """Download a generated report file."""
    report = await db.get(Report, report_id)
    if not report or not os.path.exists(report.file_path):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report file not found"
        )
        
    # Content-type configuration
    media_types = {
        "pdf": "application/pdf",
        "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "csv": "text/csv"
    }
    
    return FileResponse(
        path=report.file_path,
        media_type=media_types.get(report.format, "application/octet-stream"),
        filename=os.path.basename(report.file_path)
    )
